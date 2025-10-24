"""
Excel Exporter for conversations
Uses openpyxl for professional Excel generation with formatting
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime
from typing import List, Dict, Optional
import io


class ExcelExporter:
    """
    Professional Excel exporter for chat conversations

    Features:
    - Multiple worksheets
    - Custom styling and formatting
    - Charts and visualizations
    - Statistics summary
    - Formulas and calculations
    """

    def __init__(self):
        """Initialize Excel exporter"""
        self.wb = None
        self._setup_styles()

    def _setup_styles(self):
        """Setup common styles"""
        # Header style
        self.header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Title style
        self.title_font = Font(name='Calibri', size=16, bold=True, color='4F46E5')
        self.title_alignment = Alignment(horizontal='center', vertical='center')

        # User message style
        self.user_fill = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')

        # Assistant message style
        self.assistant_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

        # Border style
        thin_border = Side(style='thin', color='D1D5DB')
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

        # Normal text
        self.normal_font = Font(name='Calibri', size=11)
        self.normal_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    def export_conversation(
        self,
        conversation: Dict,
        messages: List[Dict],
        output_path: Optional[str] = None,
        include_statistics: bool = True,
        include_charts: bool = True
    ) -> bytes:
        """
        Export conversation to Excel

        Args:
            conversation: Conversation metadata
            messages: List of messages
            output_path: Optional file path to save Excel
            include_statistics: Include statistics sheet
            include_charts: Include visualization charts

        Returns:
            bytes: Excel file content
        """
        self.wb = Workbook()

        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])

        # Create messages sheet
        self._create_messages_sheet(conversation, messages)

        # Create statistics sheet
        if include_statistics:
            self._create_statistics_sheet(conversation, messages)

        # Create timeline sheet
        self._create_timeline_sheet(messages)

        # Add charts if requested
        if include_charts:
            self._add_charts(messages)

        # Save to buffer
        buffer = io.BytesIO()
        self.wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()

        # Save to file if path provided
        if output_path:
            self.wb.save(output_path)

        return excel_bytes

    def _create_messages_sheet(self, conversation: Dict, messages: List[Dict]):
        """Create messages worksheet"""
        ws = self.wb.create_sheet('Messages', 0)

        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"Conversation: {conversation.get('title', 'Untitled')}"
        title_cell.font = self.title_font
        title_cell.alignment = self.title_alignment
        ws.row_dimensions[1].height = 30

        # Metadata row
        ws.merge_cells('A2:F2')
        metadata_text = (
            f"Created: {conversation.get('created_at', 'N/A')} | "
            f"Model: {conversation.get('model', 'N/A')} | "
            f"Messages: {len(messages)} | "
            f"Tokens: {conversation.get('token_count', 0)}"
        )
        ws['A2'] = metadata_text
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 20

        # Headers
        headers = ['#', 'Role', 'Timestamp', 'Content', 'Tokens', 'Model']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        ws.row_dimensions[3].height = 25

        # Message data
        for idx, message in enumerate(messages, 1):
            row = idx + 3
            role = message.get('role', 'user')

            # Row data
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=role.capitalize())
            ws.cell(row=row, column=3, value=message.get('created_at', ''))
            ws.cell(row=row, column=4, value=message.get('content', ''))
            ws.cell(row=row, column=5, value=message.get('tokens', 0))
            ws.cell(row=row, column=6, value=message.get('model', ''))

            # Apply styling
            fill = self.user_fill if role == 'user' else self.assistant_fill

            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                cell.fill = fill

                if col == 4:  # Content column
                    cell.alignment = self.normal_alignment
                    ws.row_dimensions[row].height = max(30, len(cell.value) // 50 * 15)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15

        # Freeze header rows
        ws.freeze_panes = 'A4'

    def _create_statistics_sheet(self, conversation: Dict, messages: List[Dict]):
        """Create statistics worksheet"""
        ws = self.wb.create_sheet('Statistics')

        # Title
        ws.merge_cells('A1:B1')
        ws['A1'] = 'Conversation Statistics'
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.title_alignment
        ws.row_dimensions[1].height = 30

        # Calculate statistics
        total_messages = len(messages)
        user_messages = sum(1 for m in messages if m.get('role') == 'user')
        assistant_messages = sum(1 for m in messages if m.get('role') == 'assistant')
        system_messages = sum(1 for m in messages if m.get('role') == 'system')
        total_tokens = sum(m.get('tokens', 0) for m in messages)
        avg_tokens = total_tokens / total_messages if total_messages > 0 else 0

        # Statistics data
        stats = [
            ('Conversation Info', ''),
            ('Title', conversation.get('title', 'Untitled')),
            ('Created', conversation.get('created_at', 'N/A')),
            ('Updated', conversation.get('updated_at', 'N/A')),
            ('Model', conversation.get('model', 'N/A')),
            ('Provider', conversation.get('provider', 'N/A')),
            ('', ''),
            ('Message Statistics', ''),
            ('Total Messages', total_messages),
            ('User Messages', user_messages),
            ('Assistant Messages', assistant_messages),
            ('System Messages', system_messages),
            ('', ''),
            ('Token Statistics', ''),
            ('Total Tokens', total_tokens),
            ('Average Tokens/Message', f'{avg_tokens:.2f}'),
            ('User Message %', f'{user_messages/total_messages*100:.1f}%' if total_messages > 0 else '0%'),
            ('Assistant Message %', f'{assistant_messages/total_messages*100:.1f}%' if total_messages > 0 else '0%'),
            ('', ''),
            ('Export Info', ''),
            ('Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Generated By', 'AI Chat Studio')
        ]

        # Write statistics
        for row, (label, value) in enumerate(stats, 3):
            if label and not value:
                # Section header
                ws.merge_cells(f'A{row}:B{row}')
                cell = ws.cell(row=row, column=1, value=label)
                cell.font = Font(name='Calibri', size=12, bold=True, color='4F46E5')
                ws.row_dimensions[row].height = 25
            elif label:
                # Data row
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
            # Empty row - just skip

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30

    def _create_timeline_sheet(self, messages: List[Dict]):
        """Create timeline analysis sheet"""
        ws = self.wb.create_sheet('Timeline')

        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = 'Message Timeline'
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.title_alignment

        # Headers
        headers = ['Time', 'Role', 'Tokens', 'Cumulative Tokens']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment

        # Timeline data
        cumulative_tokens = 0
        for idx, message in enumerate(messages, 1):
            tokens = message.get('tokens', 0)
            cumulative_tokens += tokens

            ws.cell(row=idx+2, column=1, value=message.get('created_at', ''))
            ws.cell(row=idx+2, column=2, value=message.get('role', ''))
            ws.cell(row=idx+2, column=3, value=tokens)
            ws.cell(row=idx+2, column=4, value=cumulative_tokens)

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18

    def _add_charts(self, messages: List[Dict]):
        """Add visualization charts"""
        ws = self.wb['Statistics']

        # Message distribution pie chart
        pie = PieChart()
        pie.title = "Message Distribution"
        pie.height = 10
        pie.width = 15

        # Data for pie chart
        user_count = sum(1 for m in messages if m.get('role') == 'user')
        assistant_count = sum(1 for m in messages if m.get('role') == 'assistant')

        # Add chart data manually
        chart_data = [
            ['Type', 'Count'],
            ['User', user_count],
            ['Assistant', assistant_count]
        ]

        start_row = 25
        for row_idx, row_data in enumerate(chart_data, start_row):
            for col_idx, value in enumerate(row_data, 4):  # Column D
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Create chart
        data = Reference(ws, min_col=5, min_row=start_row, max_row=start_row+2, max_col=5)
        categories = Reference(ws, min_col=4, min_row=start_row+1, max_row=start_row+2)

        pie.add_data(data, titles_from_data=True)
        pie.set_categories(categories)

        ws.add_chart(pie, "D30")

    def export_multiple_conversations(
        self,
        conversations: List[Dict],
        output_path: str
    ) -> bytes:
        """
        Export multiple conversations to Excel with separate sheets

        Args:
            conversations: List of conversations with messages
            output_path: Path to save Excel file

        Returns:
            bytes: Excel content
        """
        self.wb = Workbook()

        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])

        # Summary sheet
        self._create_summary_sheet(conversations)

        # Individual conversation sheets
        for idx, conv_data in enumerate(conversations[:10], 1):  # Limit to 10 sheets
            conversation = conv_data.get('conversation', {})
            messages = conv_data.get('messages', [])

            ws = self.wb.create_sheet(f'Conv {idx}')
            self._create_simple_messages_sheet(ws, conversation, messages)

        # Save
        buffer = io.BytesIO()
        self.wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()

        if output_path:
            self.wb.save(output_path)

        return excel_bytes

    def _create_summary_sheet(self, conversations: List[Dict]):
        """Create summary sheet for multiple conversations"""
        ws = self.wb.create_sheet('Summary', 0)

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = 'Conversations Summary'
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.title_alignment

        # Headers
        headers = ['#', 'Title', 'Messages', 'Tokens', 'Created', 'Updated']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment

        # Data
        for idx, conv_data in enumerate(conversations, 1):
            conversation = conv_data.get('conversation', {})
            messages = conv_data.get('messages', [])

            ws.cell(row=idx+2, column=1, value=idx)
            ws.cell(row=idx+2, column=2, value=conversation.get('title', 'Untitled'))
            ws.cell(row=idx+2, column=3, value=len(messages))
            ws.cell(row=idx+2, column=4, value=sum(m.get('tokens', 0) for m in messages))
            ws.cell(row=idx+2, column=5, value=conversation.get('created_at', ''))
            ws.cell(row=idx+2, column=6, value=conversation.get('updated_at', ''))

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20

    def _create_simple_messages_sheet(self, ws, conversation: Dict, messages: List[Dict]):
        """Create simplified messages sheet"""
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = conversation.get('title', 'Untitled')
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.title_alignment

        # Headers
        headers = ['Role', 'Content', 'Tokens', 'Time']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Messages
        for idx, message in enumerate(messages, 1):
            ws.cell(row=idx+2, column=1, value=message.get('role', ''))
            ws.cell(row=idx+2, column=2, value=message.get('content', ''))
            ws.cell(row=idx+2, column=3, value=message.get('tokens', 0))
            ws.cell(row=idx+2, column=4, value=message.get('created_at', ''))

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 20
